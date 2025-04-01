import base64
import io
import logging
import openpyxl
from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

class FeedbackImportUsersExcelWizard(models.TransientModel):
    _name = 'feedback.import.users.excel.wizard'
    _description = 'Wizard to import Users from Excel'

    file_data = fields.Binary("Excel File", required=True)
    file_name = fields.Char("File Name")

    def action_import_users(self):
        """Read the Excel file and create users based on its data."""
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        try:
            # Decode the uploaded file and load the workbook
            file_content = base64.b64decode(self.file_data)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file_content))
            sheet = workbook.active

            # Assume the first row contains headers:
            headers = [cell.value for cell in sheet[1]]
            # Expected headers for users: 
            # firstname, lastname, email, phone, country, dob, job, role, total_time_spent, idd, token, password
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                self.env['users.users'].create({
                    'firstname': data.get('firstname'),
                    'lastname': data.get('lastname'),
                    'email': data.get('email'),
                    'phone': data.get('phone'),
                    'country': data.get('country'),
                    'dob': data.get('dob') if data.get('dob') else False,
                    'job': data.get('job'),
                    'role': data.get('role'),
                    'total_time_spent': data.get('total_time_spent') or 0.0,
                    'idd': data.get('idd'),
                    'token': data.get('token'),
                    'password': data.get('password'),
                })
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': _('Users imported successfully.'),
                    'type': 'success',
                    'sticky': False,
                }
            }
        except Exception as e:
            _logger.exception("Error importing Users from Excel")
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Error'),
                    'message': _('Failed to import Users: %s') % e,
                    'type': 'danger',
                    'sticky': False,
                }
            }


class FeedbackImportFeedbackExcelWizard(models.TransientModel):
    _name = 'feedback.import.feedback.excel.wizard'
    _description = 'Wizard to import Feedbacks from Excel'

    file_data = fields.Binary("Excel File", required=True)
    file_name = fields.Char("File Name")

    def action_import_feedbacks(self):
        """Read the Excel file and create feedback records based on its data."""
        if not self.file_data:
            raise UserError(_("Please upload an Excel file."))

        try:
            file_content = base64.b64decode(self.file_data)
            workbook = openpyxl.load_workbook(filename=io.BytesIO(file_content))
            sheet = workbook.active

            # Assume the first row contains headers:
            headers = [cell.value for cell in sheet[1]]
            # Expected headers for feedbacks: 
            # feedback, user_email, property_id, created_at
            for row in sheet.iter_rows(min_row=2, values_only=True):
                data = dict(zip(headers, row))
                user = self.env['users.users'].search([('email', '=', data.get('user_email'))], limit=1)
                if not user:
                    _logger.warning("User with email %s not found; skipping row.", data.get('user_email'))
                    continue
                if not data.get('property_id'):
                    _logger.warning("No property_id provided; skipping row.")
                    continue
                self.env['real.estate.feedback'].create({
                    'feedback': data.get('feedback'),
                    'user_id': user.id,
                    'property_id': data.get('property_id'),
                    'created_at': data.get('created_at') if data.get('created_at') else fields.Datetime.now(),
                })
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Success'),
                    'message': _('Feedbacks imported successfully.'),
                    'type': 'success',
                    'sticky': False,
                }
            }
        except Exception as e:
            _logger.exception("Error importing Feedbacks from Excel")
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Error'),
                    'message': _('Failed to import Feedbacks: %s') % e,
                    'type': 'danger',
                    'sticky': False,
                }
            }
